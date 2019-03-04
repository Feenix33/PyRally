#!python3
from __future__ import print_function
import os
import sys
import csv
import getopt
import datetime
import argparse

from pyral import Rally, rallySettings, rallyWorkset
from openpyxl import Workbook

"""
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
Features to add
   add tags
   remove config and use args only?

Missing Error handling
   search not found
   files not opening
   unknown flags
   upper/lowercase
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
"""


gFieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature.FormattedID', 'TeamFeature.Name',
                        'ScheduleState', 'Project.Name', 'Iteration.Name', 'Owner.Name', 'CreationDate', 'Tags']
gPPMFieldnames = ['FormattedID', 'Name', 'LeafStoryPlanEstimateTotal', 'Parent.FormattedID', 'Parent.Name', 
                        'LeafStoryCount', 'AcceptedLeafStoryPlanEstimateTotal', 'AcceptedLeafStoryCount', 'Owner.Name', 'CreationDate']



class theConfig:
    def __init__(self):
        self.outToConsole = True

        self.writeCSV = True
        self.writeExcel = True
        self.fnameCSV = "out.csv"
        self.fnameExcel = "out.xlsx"
        self.fnameExceloutType = 0
        self.singleExcel = True
        self.fnameInput = 'in.txt'
        self.strNoExist = "--"
        self.useInput = True
        self.logging = True
        self.logFile = None
        self.logFilename = 'log.txt'
        self.currentIteration = False
        self.previousIteration = False
        self.addDate = False


gConfig = theConfig()

# converting to current for 2018
# int( (target - 20dec2017) /14)
def convertIterIDtoLabel2018(iterID):
    today = datetime.date.today()
    magicDate = datetime.date(2017, 12, 20)

    #check if current iteration
    if iterID == 'Now':
        iterID = ("S{:02d}").format (int (((today - magicDate).days) / 14))

    #check if previous iteration
    if iterID == 'Pre':
        iterID = ("S{:02d}").format (int (((today - datetime.timedelta(days=14) - magicDate).days) / 14))

    # convert the short version of the iteration ID to the full version
    elif len(iterID) != 3 and len(iterID) != 8:
        return iterID

    # modify for 2019
    iterNum = int(iterID[1:3])
    iterNumNew = ((iterNum-1) % 26) + 1
    iterID = ("S{:02d}").format(iterNumNew) + iterID[4:]

    if len(iterID) == 3: iterYear = today.year
    else: iterYear = int(iterID[4:])

    iter1start = datetime.date(2017, 1, 5)
    iter1end = datetime.date(2017, 1, 18)
    deltaDays = (iterYear - 2017) * 364 + (iterNum -1)*14
    iterStart = iter1start + datetime.timedelta(deltaDays)
    iterEnd = iterStart + datetime.timedelta(14)
    strIterLabel = ("S{:02d}#{:04d}-{:02d}-{:02d}/{:02d}-{:02d}").format(
            iterNumNew, iterYear, iterStart.month, iterStart.day, iterEnd.month, iterEnd.day)
    return strIterLabel

def convertIterIDtoLabel(iterID):
    # This is a hack. Only for 2019 and fails on S01 previous
    today = datetime.date.today()
    magicDate = datetime.date(2019,  1, 2) #S01 for 2019
    iterYear = 2019

    if iterID == 'Now':
        iterNum = (int (((today - magicDate).days) / 14)) + 1
    elif iterID == 'Pre':
        iterNum = (int (((today - magicDate).days) / 14))
    elif len(iterID) != 3 and len(iterID) != 8:
        return iterID

    iterStart = magicDate + (iterNum-1)*datetime.timedelta(14)
    iterEnd = iterStart + datetime.timedelta(13)
    strIterLabel = ("S{:02d}#{:04d}-{:02d}-{:02d}/{:02d}-{:02d}").format(
            iterNum, iterYear, iterStart.month, iterStart.day, iterEnd.month, iterEnd.day)
    return strIterLabel

def processRecord(record):
    return  [getattrError(record, attr) for attr in gFieldnames] 

 
def getattrError(n, attr):
    if attr == 'Tags':
        return ('|'.join([i.Name for i in n.Tags]))
    try:
        return getattr(n, attr)
    except AttributeError:
        dot = attr.find(".")
        if dot == -1 :
            return str(gConfig.strNoExist)
        try:
            return getattr( getattr(n, attr[:dot]), attr[(dot+1):])
        except AttributeError:
            return str(gConfig.strNoExist)

def processIteration(instRally, nameIter, queryToken):
    records = []
    entityList = ['HierarchicalRequirement', 'Defect']
    #entityName = 'Defect'
    fullIterationName = convertIterIDtoLabel(nameIter[2:])
    queryString = 'Iteration.Name = "' + fullIterationName + '"'

    for entityName in entityList:
        response = instRally.get(entityName, fetch=True, projectScopeDown=True, query=queryString)
        #response = gRally.get(queryItem, fetch=True, projectScopeDown=True, query=gblIteration) 

        consoleStatus ("   " + queryString + "    Count = " + str(response.resultCount))
        for story in response:
            records.append( processRecord( story ) + [queryToken[2:]]) #xx
    return records

def processFEA(instRally, feaID, queryToken):
    entityName = 'PortfolioItem'
    queryString = 'FormattedID = "' + feaID + '"'
    response = instRally.get(entityName, fetch=True, projectScopeUp=True, query=queryString)
    records = []
    for ppmFeature in response:
        records.append( [getattrError(ppmFeature, attr) for attr in gPPMFieldnames] + [queryToken])
        if ppmFeature.DirectChildrenCount > 0 and hasattr(ppmFeature, 'Children'):
            for teamfeature in ppmFeature.Children:
                # add the team feature to the data table
                records.append( [getattrError(teamfeature, attr) for attr in gPPMFieldnames] + [queryToken])
                # add the subsequent children
                records.extend( processStory(instRally, teamfeature.FormattedID, queryToken))
    return records 

def processStory(instRally, storyID, queryToken): # story is user story or defect or TF
    records = []
    if storyID[0] == "U":
        entityName = 'HierarchicalRequirement'
        queryString = 'FormattedID = "' + storyID + '"'
    elif storyID[0] == "D":
        entityName = 'Defect'
        queryString = 'FormattedID = "' + storyID + '"'
    else:
        entityName = 'HierarchicalRequirement'
        queryString = 'Feature.FormattedID = "' + storyID + '"'
    response = instRally.get(entityName, fetch=True, projectScopeDown=True, query=queryString)
    consoleStatus ("   " + queryString + "    Count = " + str(response.resultCount))
    for story in response:
        records.append( processRecord( story ) + [queryToken]) #xx
    return records


def tokens(fileobj):
    for line in fileobj:
        if line[0] != '#':
            for word in line.split():
                yield word

def getHeadList(ppmType=None):
    #return the header items to write in csv or xlsx
    if ppmType: return (gPPMFieldnames + ['QueryToken'])
    else: return (gFieldnames + ['QueryToken'])

def writeCSV(listData, fname):
    fout = open(fname, 'wb')
    fwriter = csv.writer(fout, dialect='excel')
    fwriter.writerow( getHeadList() )
    fwriter.writerows(listData)
    fout.close()

def writeExcel(listData, fname, singleSheet):
    #check if no data
    if len(listData) == 0 or len(listData[0]) == 0:
        consoleStatus("No data to write")
        return

    consoleStatus("Writing Excel")
    workbook = Workbook()
    ws = workbook.active
    queryTokenOld = "0None"
    queryToken    = listData[0][-1]

    if singleSheet:
        ws = workbook.create_sheet("Output")
        ws.append(gFieldnames + ['QueryToken'])

    for record in listData:
        queryToken = record[-1]
        if queryToken != queryTokenOld:
            #if (not ((queryToken[0] == 'U' or queryToken[0] == 'D') and (queryTokenOld[0] == 'U' or queryTokenOld[0] == 'D'))):
            if not (queryTokenOld[0] == 'U' or queryTokenOld[0] == 'D'):
                if not singleSheet:
                    sheetTitle = queryToken.replace('/','|') #excel no like slash
                    #error need to check for duplicated sheet name
                    ws = workbook.create_sheet(sheetTitle)
                    queryTokenOld = queryToken
                    if queryToken[0] == 'T' or queryToken[0] == 'F':
                        ws.append( getHeadList(True) )
                    else:
                        ws.append( getHeadList() )
        ws.append(record)
    #get rid of the initial sheet
    #std = workbook.get_sheet_by_name('Sheet') #changing for deprecated function call
    #workbook.remove_sheet(std)
    std = workbook['Sheet']
    workbook.remove(std)
    workbook.save(fname)

def logOpen():
    if gConfig.logging: gConfig.logFile = open(gConfig.logFilename, 'w')

def log(message):
    if gConfig.logging: gConfig.logFile.write (message)

def logClose():
    if gConfig.logFile: gConfig.logFile.close()

def consoleStatus(message):
    if gConfig.outToConsole: print (message)

def addDateFilename(inName):
    newName = inName
    if len(newName) > 5: #magic for .xlsx
        if newName[-5:] == '.xlsx':
            newName = newName[:-5]
    today = datetime.date.today()

    newName = newName + (".{:04d}.{:02d}.{:02d}.xlsx").format(
            today.year, today.month, today.day)
    return newName

def buildInputParser(parser):
    parser.add_argument("-i", "--infile", type=argparse.FileType('r'), default='in.txt',
            help="input file with search parameters")
    parser.add_argument("-q", "--quiet", action="store_true", help="No console messages")
    parser.add_argument("-c", "--csv", nargs='?', const=1, type=str, default='', help="Write a .csv output, default=out.csv")
    parser.add_argument("-x", "-xl", "--excel", nargs='?', const=1, type=str, default='out.xlsx', help="Write a Excel output")
    parser.add_argument("-l", "--log", action="store_true", help="Turn on logging")
    parser.add_argument("-d", "--daily", action="store_true", help="Do the daily extract")
    parser.add_argument("-nx", "--noxl", action='store_true', default=False, help="Suppress Excel output")
    parser.add_argument("-na", "--nastring", nargs=1, type=str, default="xx", help="String for empty, default='xx'")
    parser.add_argument("-x1", "--xlone", action='store_true', default=False, help="Excel file output as single sheet")
    parser.add_argument("-it", "--iter", action='store_true', default=False, help="Just do the current iteration")
    parser.add_argument("-i2", "--iter2", action='store_true', default=False, help="Do the current and previous iteration")
    parser.add_argument("-ad", "--adddate", action='store_true', default=False, help="Add the date to the filename")
    parser.add_argument('baz',default=[], nargs='*', help="Parameters to search for")

def mapArgsserToGlobal(args):
    #convert the passed arguments to the global configuration and return the search tokens
    gConfig.outToConsole = not args.quiet
    gConfig.fnameInput = args.infile.name
    gConfig.writeExcel = True
    gConfig.logging = args.log # logging control

    if args.daily:
        args.iter = True
        args.adddate = True
        args.excel  = "pyRallyLGS.xlsx"

    if args.excel == 1: gConfig.fnameExcel = "out.xlsx"
    else: 
        gConfig.fnameExcel = args.excel
        if gConfig.fnameExcel[-5:] != '.xlsx': gConfig.fnameExcel += '.xlsx'

    gConfig.writeCSV = True
    if args.csv == "": gConfig.writeCSV = False
    elif args.csv == 1: gConfig.fnameCSV = "out.csv"
    else: 
        gConfig.fnameCSV = args.csv
        if gConfig.fnameCSV[-4:] != '.csv': gConfig.fnameCSV += '.csv'

    if args.noxl: gConfig.writeExcel = False

    gConfig.singleExcel = args.xlone

    gConfig.strNoExist = args.nastring[0]

    if args.adddate:
        gConfig.addDate = True
        gConfig.fnameExcel = addDateFilename(gConfig.fnameExcel)

    if args.iter or args.iter2: # we are just going to do the current iteration, cancel the other inputs
        gConfig.currentIteration = True
        gConfig.useInput = False
        args.baz = ['ITNow']
        if args.iter2:
            gConfig.previousIteration = True
            args.baz.append('ITPre')

    if len(args.baz) > 0: gConfig.useInput = False
    return args.baz


def main():
    dataHR = []
    queryList = []

    inParser = argparse.ArgumentParser()
    buildInputParser(inParser)
    args = inParser.parse_args()
    queryList.extend(mapArgsserToGlobal(args))

    logOpen()

    attrs = vars(gConfig)
    for item in attrs.items(): log( "CONFIG %s: %s\n" % item)
    log("-"*60+'\n')

    finputfile = gConfig.fnameInput
    if gConfig.useInput:
        consoleStatus('Getting search tokens...')
        try:
            infile = open(finputfile, 'r')
            queryList.extend(tokens(infile))

        except IOError:
            print ("IOError: Cannot open", sys.argv[0], "<input file>")

    for queryItem in queryList: log('QUERY. {0}\n'.format(str(queryItem)))
    log("-"*60+'\n')

    consoleStatus('Logging in...')
    rally = Rally(server, apikey=apikey, workspace=workspace, project=project)

    consoleStatus('Query execution...')
    for queryItem in queryList:
        log("QUERY: %s\n" % queryItem)
        if queryItem[:2] == "FE":
            dataHR.extend (processFEA(rally, queryItem, queryItem))
        elif queryItem[:2] == "PR":
            dataHR.extend (processFEA(rally, queryItem, queryItem))
        elif queryItem[:2] == "TF":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "US":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "DE":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "IT":
            dataHR.extend (processIteration(rally, queryItem, queryItem))
        else:
            print ("Error query for " + queryItem)


    if gConfig.writeCSV: writeCSV(dataHR, gConfig.fnameCSV)
    if gConfig.writeExcel: writeExcel(dataHR, gConfig.fnameExcel, gConfig.singleExcel)

    consoleStatus('Fini')
    logClose()

def test_iter():
    """
    Test the iteration string creation because it is based on magic formulas
    """
    print ("Now", convertIterIDtoLabel("Now"))
    print ("Pre", convertIterIDtoLabel("Pre"))
    #d1 = datetime.date(2019, 2, 13)
    #d2 = datetime.date(2019, 2, 26)
    #print (d1, d2, (d2-d1).days)
    #d1 = d2 + datetime.timedelta(days=1) 
    #d2 = datetime.date(2019, 3, 12)
    #print (d1, d2, (d2-d1).days)


if __name__ == '__main__':
    #test_iter()
    server = 'rally1.rallydev.com'
    apikey = '_LhzUHJ1GQJQWkEYepqIJV9NO96FkErDpQvmHG4WQ'
    workspace = 'Sabre Production Workspace'
    project = 'LGS' 
    main()

